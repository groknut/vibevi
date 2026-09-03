import json
import os
import xml.etree.ElementTree as ET
from pathlib import Path


def _file_meta(path: str) -> dict:
    """Collect common file metadata.

    Returns:
        dict: {
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    size = os.path.getsize(path)
    return {"path": path, "name": os.path.basename(path), "size": size}


def parse_txt(path: str) -> dict:
    """Parse a plain text file.

    Args:
        path: str — path to the text file.

    Returns:
        dict: {
            "type": str — always "text",
            "content": str — full file content,
            "lines": int — number of lines,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    lines = text.splitlines()
    return {
        "type": "text",
        "content": text,
        "lines": len(lines),
        **_file_meta(path),
    }


def parse_md(path: str) -> dict:
    """Parse a Markdown file.

    Args:
        path: str — path to the Markdown file.

    Returns:
        dict: {
            "type": str — always "markdown",
            "content": str — raw Markdown text,
            "lines": int — number of lines,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    result = parse_txt(path)
    result["type"] = "markdown"
    return result


def parse_log(path: str) -> dict:
    """Parse a log file.

    Args:
        path: str — path to the log file.

    Returns:
        dict: {
            "type": str — always "log",
            "content": str — full log content,
            "lines": int — number of lines,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    result = parse_txt(path)
    result["type"] = "log"
    return result


def parse_json(path: str) -> dict:
    """Parse a JSON file.

    Args:
        path: str — path to the JSON file.

    Returns:
        dict: {
            "type": str — always "json",
            "content": str — pretty-printed JSON string,
            "content_type": str — Python type name ("dict", "list", etc.),
            "keys": list[str] — top-level keys (if object),
            "items": int — number of items (if dict or list),
            "value": Any — raw value (if scalar),
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    formatted = json.dumps(data, indent=2, ensure_ascii=False)
    content_type = type(data).__name__

    meta = {"type": "json", "content": formatted, "content_type": content_type}

    if isinstance(data, dict):
        meta["keys"] = list(data.keys())
        meta["items"] = len(data)
    elif isinstance(data, list):
        meta["items"] = len(data)
    else:
        meta["value"] = data

    meta.update(_file_meta(path))
    return meta


def _dump_element(element: ET.Element, lines: list[str], indent: int) -> None:
    """Recursively serialize an XML element tree into indented string lines."""
    prefix = "  " * indent
    attrs = " ".join(f'{k}="{v}"' for k, v in element.attrib.items())
    tag = element.tag
    text = (element.text or "").strip()

    if attrs:
        tag = f"{tag} {attrs}"

    if len(element) == 0 and text:
        lines.append(f"{prefix}<{tag}> {text} </{element.tag}>")
    elif len(element) == 0:
        lines.append(f"{prefix}<{tag} />")
    else:
        lines.append(f"{prefix}<{tag}>")
        if text:
            lines.append(f"{prefix}  {text}")
        for child in element:
            _dump_element(child, lines, indent + 1)
        lines.append(f"{prefix}</{element.tag}>")


def parse_xml(path: str) -> dict:
    """Parse an XML file.

    Args:
        path: str — path to the XML file.

    Returns:
        dict: {
            "type": str — always "xml",
            "content": str — indented XML string,
            "root_tag": str — tag name of the root element,
            "elements": int — total number of elements in tree,
            "attributes": dict — attributes of the root element,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    tree = ET.parse(path)
    root = tree.getroot()
    lines = []
    _dump_element(root, lines, indent=0)

    element_count = sum(1 for _ in root.iter())
    return {
        "type": "xml",
        "content": "\n".join(lines),
        "root_tag": root.tag,
        "elements": element_count,
        "attributes": dict(root.attrib),
        **_file_meta(path),
    }


def parse_pdf(path: str) -> dict:
    """Parse a PDF file using pdfplumber.

    Args:
        path: str — path to the PDF file.

    Returns:
        dict: {
            "type": str — always "pdf",
            "content": str — all pages joined by double newline,
            "pages": int — number of pages,
            "page_texts": list[dict] — per-page data: {"page": int, "text": str},
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    try:
        import pdfplumber
    except ImportError:
        return {"type": "pdf", "content": "[pdfplumber not installed]", **_file_meta(path)}

    pages = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            pages.append({"page": i + 1, "text": text})

    full_text = "\n\n".join(p["text"] for p in pages)
    return {
        "type": "pdf",
        "content": full_text,
        "pages": len(pages),
        "page_texts": pages,
        **_file_meta(path),
    }


def parse_docx(path: str) -> dict:
    """Parse a DOCX file using python-docx.

    Args:
        path: str — path to the DOCX file.

    Returns:
        dict: {
            "type": str — always "docx",
            "content": str — all paragraphs joined by newline,
            "paragraphs": int — number of paragraphs,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    try:
        import docx
    except ImportError:
        return {"type": "docx", "content": "[python-docx not installed]", **_file_meta(path)}

    doc = docx.Document(path)
    paragraphs = [p.text for p in doc.paragraphs]
    text = "\n".join(paragraphs)
    return {
        "type": "docx",
        "content": text,
        "paragraphs": len(paragraphs),
        **_file_meta(path),
    }


def parse_xlsx(path: str) -> dict:
    """Parse an XLSX file using openpyxl.

    Args:
        path: str — path to the XLSX file.

    Returns:
        dict: {
            "type": str — always "xlsx",
            "content": str — all sheets formatted as text,
            "sheet_names": list[str] — list of sheet names,
            "sheets": dict[str, list[list[str]]] — sheet name to rows mapping,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"type": "xlsx", "content": "[openpyxl not installed]", **_file_meta(path)}

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(cells)
        sheets[sheet_name] = rows
    wb.close()

    lines = []
    for name, rows in sheets.items():
        lines.append(f"=== Sheet: {name} ===")
        for row in rows:
            lines.append("\t".join(row))

    return {
        "type": "xlsx",
        "content": "\n".join(lines),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
        **_file_meta(path),
    }


def parse_xls(path: str) -> dict:
    """Parse an XLS file using xlrd.

    Args:
        path: str — path to the XLS file.

    Returns:
        dict: {
            "type": str — always "xls",
            "content": str — all sheets formatted as text,
            "sheet_names": list[str] — list of sheet names,
            "sheets": dict[str, list[list[str]]] — sheet name to rows mapping,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    try:
        import xlrd
    except ImportError:
        return {"type": "xls", "content": "[xlrd not installed]", **_file_meta(path)}

    wb = xlrd.open_workbook(path)
    sheets = {}
    for sheet in wb.sheets():
        rows = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            rows.append(cells)
        sheets[sheet.name] = rows

    lines = []
    for name, rows in sheets.items():
        lines.append(f"=== Sheet: {name} ===")
        for row in rows:
            lines.append("\t".join(row))

    return {
        "type": "xls",
        "content": "\n".join(lines),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
        **_file_meta(path),
    }


def parse_image(path: str) -> dict:
    """Parse an image file using Pillow.

    Args:
        path: str — path to the image file.

    Returns:
        dict: {
            "type": str — always "image",
            "content": str — human-readable image description,
            "format": str — image format (PNG, JPEG, etc.),
            "width": int — image width in pixels,
            "height": int — image height in pixels,
            "mode": str — color mode (RGB, RGBA, L, etc.),
            "info": dict[str, str] — additional image metadata,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    try:
        from PIL import Image
    except ImportError:
        return {"type": "image", "content": "[Pillow not installed]", **_file_meta(path)}

    img = Image.open(path)
    w, h = img.size
    fmt = img.format or "unknown"
    mode = img.mode

    info = {}
    for k, v in img.info.items():
        if isinstance(v, (str, int, float)):
            info[k] = str(v)

    img.close()

    description_lines = [
        f"Format: {fmt}",
        f"Size: {w} x {h}",
        f"Mode: {mode}",
    ]
    for k, v in info.items():
        description_lines.append(f"{k}: {v}")

    return {
        "type": "image",
        "content": "\n".join(description_lines),
        "format": fmt,
        "width": w,
        "height": h,
        "mode": mode,
        "info": info,
        **_file_meta(path),
    }


def parse_raw(path: str) -> dict:
    """Read any file as raw bytes for unknown or unsupported types.

    Args:
        path: str — path to any file.

    Returns:
        dict: {
            "type": str — always "raw",
            "content": str — hex dump preview of the first bytes,
            "data": bytes — full raw file content,
            "path": str — full file path,
            "name": str — file name only,
            "size": int — file size in bytes,
        }
    """
    with open(path, "rb") as f:
        data = f.read()

    preview_size = min(len(data), 512)
    hex_lines = []
    for offset in range(0, preview_size, 16):
        chunk = data[offset:offset + 16]
        hex_part = " ".join(f"{b:02x}" for b in chunk)
        ascii_part = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
        hex_lines.append(f"{offset:08x}  {hex_part:<48s}  {ascii_part}")

    return {
        "type": "raw",
        "content": "\n".join(hex_lines),
        "data": data,
        **_file_meta(path),
    }


PARSERS = {
    ".txt": parse_txt,
    ".md": parse_md,
    ".log": parse_log,
    ".json": parse_json,
    ".xml": parse_xml,
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xls,
    ".png": parse_image,
    ".jpeg": parse_image,
    ".jpg": parse_image,
}


def parse_file(path: str) -> dict:
    """Dispatch to the appropriate parser by file extension.

    For unsupported extensions, falls back to parse_raw which reads the file
    as raw bytes and returns a hex dump preview.

    Args:
        path: str — path to any file.

    Returns:
        dict — result from the matched parser. Always contains "type", "content",
        "path", "name", "size". Additional keys depend on the file type.
        On unsupported extension returns type="raw", on error returns type="error".
    """
    ext = Path(path).suffix.lower()
    parser = PARSERS.get(ext)
    if parser is None:
        try:
            return parse_raw(path)
        except Exception as e:
            return {
                "type": "error",
                "content": f"[Error reading raw file]: {e}",
                **_file_meta(path),
            }
    try:
        return parser(path)
    except Exception as e:
        return {
            "type": "error",
            "content": f"[Error parsing {ext}]: {e}",
            **_file_meta(path),
        }
