from typing import TypedDict
from ._meta import FileMeta, file_meta


class SpreadsheetResult(FileMeta):
    type: str
    content: str
    sheet_names: list[str]
    sheets: dict[str, list[list[str]]]


def _escape_html(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _format_sheets(sheets: dict[str, list[list[str]]]) -> str:
    """Format sheets dict into HTML tables."""
    parts: list[str] = []
    for name, rows in sheets.items():
        escaped_name = _escape_html(name)
        parts.append(f'<h3 style="margin: 16px 0 8px;">Sheet: {escaped_name}</h3>')
        parts.append('<table border="1" cellpadding="4" cellspacing="0" '
                     'style="border-collapse: collapse; width: 100%;">')
        if rows:
            header = rows[0]
            parts.append('<thead><tr style="background: #f0f0f0; font-weight: bold;">')
            for cell in header:
                parts.append(f"<td>{_escape_html(cell)}</td>")
            parts.append("</tr></thead>")
            parts.append("<tbody>")
            for row in rows[1:]:
                parts.append("<tr>")
                for cell in row:
                    parts.append(f"<td>{_escape_html(cell)}</td>")
                parts.append("</tr>")
            parts.append("</tbody>")
        parts.append("</table>")
    return f'<div style="padding: 10px; font-family: sans-serif;">{"".join(parts)}</div>'


def parse_xlsx(path: str) -> SpreadsheetResult:
    """Parse an XLSX file using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"type": "xlsx", "content": "[openpyxl not installed]", "sheet_names": [], "sheets": {}, **file_meta(path)}

    wb = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[list[str]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(cells)
        sheets[sheet_name] = rows
    wb.close()

    return {
        "type": "xlsx",
        "content": _format_sheets(sheets),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
        **file_meta(path),
    }


def parse_xls(path: str) -> SpreadsheetResult:
    """Parse an XLS file using xlrd."""
    try:
        import xlrd
    except ImportError:
        return {"type": "xls", "content": "[xlrd not installed]", "sheet_names": [], "sheets": {}, **file_meta(path)}

    wb = xlrd.open_workbook(path)
    sheets: dict[str, list[list[str]]] = {}
    for sheet in wb.sheets():
        rows: list[list[str]] = []
        for row_idx in range(sheet.nrows):
            cells = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
            rows.append(cells)
        sheets[sheet.name] = rows

    return {
        "type": "xls",
        "content": _format_sheets(sheets),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
        **file_meta(path),
    }
